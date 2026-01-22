import os
import re
import pandas as pd
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
import requests

# Helper: calculate working days between two dates
def working_days_between(start, end):
    days = 0
    current = start
    while current <= end:
        if current.weekday() < 5:  # 0=Mon, 6=Sun
            days += 1
        current += timedelta(days=1)
    return days

def get_ipos():
    today = datetime.today().date()

    options = Options()
    options.add_argument("--headless")
    # run without UI
    options.add_argument("--disable-gpu")
    # disable GPU acceleration
    options.add_argument("--window-size=1920,1080")
    # set window size for rendering
    options.add_argument("--no-sandbox")
    # optional, useful in some environments
    options.add_argument("--disable-dev-shm-usage")
    # optional, prevents resource issues

    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 30)
    driver.get("https://www.investorgain.com/report/live-ipo-gmp/331/all/")

    table = wait.until(EC.presence_of_element_located((By.ID, "report_table")))
    rows = table.find_elements(By.TAG_NAME, "tr")

    ipo_data = []
    for row in rows[1:]:  # skip header
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) > 8:
            name = cols[0].text.strip()
            gmp_text = cols[1].text.strip()
            sub = cols[3].text.strip()
            start = cols[7].text.strip()
            end = cols[8].text.strip()

            # Extract GMP percentage from text like "₹12.8 (55.65%)"
            match = re.search(r"\(([\d\.]+)%\)", gmp_text)
            gmp_value = float(match.group(1)) if match else 0

            # Convert dates like "16-Jan" to datetime.date with current year
            try:
                # Extract only the date part (e.g., "13-Jan") from strings like "13-Jan GMP: 20"
                def extract_date(text, today):
                    match = re.search(r"\d{1,2}-[A-Za-z]{3}", text)
                    if match:
                        return datetime.strptime(match.group(), "%d-%b").date().replace(year=today.year)
                    return None

                start_date = extract_date(start, today)
                end_date = extract_date(end, today)

            except:
                continue

            ipo_data.append((name, gmp_value, start_date, end_date, sub))
    driver.quit()
    return ipo_data

def send_telegram_message(message):
    # Telegram config
    TELEGRAM_TOKEN = os.getenv("TG_BOT_TOKEN")
    TELEGRAM_CHAT_ID = os.getenv("TG_CHAT_ID")

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message}
    try:
        requests.post(url, data=payload)
    except Exception as e:
        print("Telegram send failed:", e)

def update_excel(ipos):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "TestData", "IPO_GMP.xlsx")
    print(excel_file)

    today = datetime.today().date()
    df = pd.read_excel(excel_file)

    # --- Add or update IPOs ---
    for name, gmp, start, end, sub in ipos:
        if not end:
            continue

        # Future IPOs (≥3 working days, within 60 days)
        if end > today and (end - today).days <= 60 and working_days_between(today, end) >= 3 and gmp > 12:
            existing = df[df["IPO Name"] == name]
            if not existing.empty:
                df.loc[df["IPO Name"] == name, "GMP"] = str(existing["GMP"].values[0]) + f",{gmp}"
            else:
                df = pd.concat([df, pd.DataFrame([[name, gmp, start, end, sub, ""]],
                                                 columns=df.columns)], ignore_index=True)

        # Closing day itself (always append GMP)
        elif end == today and (end - today).days <= 60:
            existing = df[df["IPO Name"] == name]
            if not existing.empty:
                df.loc[df["IPO Name"] == name, "GMP"] = str(existing["GMP"].values[0]) + f",{gmp}"

    # --- Evaluate status only on closing day or day before ---
    for idx, row in df.iterrows():
        end_date = pd.to_datetime(row["End Date"]).date()
        if today == end_date or today == end_date - timedelta(days=1):
            gmp_values = [float(x) for x in str(row["GMP"]).split(",") if x]
            if gmp_values:
                avg_gmp = sum(gmp_values)/len(gmp_values)
                if avg_gmp > 10:
                    df.at[idx, "Status"] = f"Proceed"
                    # --- Trigger Telegram ---
                    message = (
                        f"🚀 IPO Alert!\n\n"
                        f"Name: {row['IPO Name']}\n"
                        f"GMP History: {row['GMP']}\n"
                        f"Subscription: {row['Subscription']}\n"
                        f"Start: {row['Start Date']}, "
                        f"End: {row['End Date']}\n"
                        f"Status: Proceed"
                    )
                    send_telegram_message(message)
                else:
                    df.at[idx, "Status"] = "Skip"

    # --- Cleanup expired IPOs ---
    df = df[pd.to_datetime(df["End Date"]).dt.date >= today]

    # Save back to Excel
    df.to_excel(excel_file, index=False)

    # ---- Auto-formatting ----
    wb = load_workbook(excel_file)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(excel_file)


# Run daily
ipos = get_ipos()
update_excel(ipos)
